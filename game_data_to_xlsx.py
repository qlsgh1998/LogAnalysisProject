def write_game_data_xlsx(user_name):
    from openpyxl import load_workbook

    data_file = load_workbook('_files/' + user_name + '/data.xlsx')

    outline_sheet = data_file['outline']
    store_sheet = data_file['store']
    customer_sheet = data_file['customer']
    recipe_sheet = data_file['recipe']
    quest_sheet = data_file['quest']

    # 파일 열고 읽기
    with open('_files/' + user_name + '/Log - ' + user_name + '.txt', encoding='utf-16') as f:
        lines = f.readlines()

    # 셀 내용 채우기 함수 실행
    fill_outline_sheet(outline_sheet, lines)
    fill_store_sheet(store_sheet, lines)
    fill_customer_sheet(customer_sheet, lines)
    fill_recipe_sheet(recipe_sheet, lines)
    fill_quest_sheet(quest_sheet, lines)

    # 파일 저장 및 닫기
    data_file.save(filename='_files/' + user_name + '/data.xlsx')
    data_file.close()

    # 끝났는지 확인
    print(user_name, 'game_data done')


# 시트 0(outline) 데이터를 채움
def fill_outline_sheet(sheet, lines):
    daily_data = []  # 로그를 일자별로 나눈 것
    cut_line_by_day(daily_data, lines)
    # print(daily_data[0])

    # Day 시작 로그에서 골드, 다이아 찾아서 채우기
    # 마감 시 로그에서 방문 손님 수, 일간 순이익 찾아서 채우기
    # 매장 영업 중 로그에서 레벨업, 스킬사용, 포션사용 찾아서 채우기

    draw_data = 'ㅇ'  # 빈 칸 채우는 동그라미 모양
    level_up_index = [1]  # 레벨업 한 날짜 인덱스 리스트

    # 이 곳에서의 인덱스는 동일하게 2
    a_index = 2  # 골드, 다이아
    b_index = 2  # 방문 손님 수, 일간 다이아
    c_index = 2  # 피버 타임, 러쉬 타임, 더블 타임, 광고
    d_index = 2  # 물약 사용
    for i in range(len(daily_data)):
        # 일간 데이터
        ad_count = 0  # 광고 시청 총합
        fever_count = 0  # fever_time 스킬 사용
        rush_count = 0  # rush_time 스킬 사용
        double_count = 0  # double_time 스킬 사용
        gold_count = 0  # store_scene 골드 획득
        dia_count = 0  # store_scene 다이아 획득
        newspaper_count = 0  # before_scene newspaper 랜덤 보상 획득
        is_random_potion_counted = False  # 랜덤 물약 획득 시 여러 줄 뜨는 것을 한 번만 인식하도록 처리
        recipe_count = 0  # closed_scene recipe 기간 단축
        for j in range(len(daily_data[i])):
            # 골드, 다이아
            if 'Day' in daily_data[i][j] and '마감' not in daily_data[i][j]:
                temp_gold = daily_data[i][j + 1].split(':')[3].split('|')[0].split()
                temp_dia = daily_data[i][j + 1].split(':')[4].split()
                # print(temp_gold, ' ', temp_dia)
                sheet.cell(row=a_index, column=2).value = float(temp_gold[0])
                sheet.cell(row=a_index, column=3).value = int(temp_dia[0])
                a_index += 1
            # 방문 손님 수, 일간 순이익
            if '마감' in daily_data[i][j]:
                temp_customer = daily_data[i][j + 1].split(':')[1].split('|')[0].split()
                temp_net_income = daily_data[i][j + 1].split(':')[5].split()
                # print(temp_customer, ' ', temp_net_income)
                sheet.cell(row=b_index, column=4).value = int(temp_customer[0])
                sheet.cell(row=b_index, column=5).value = float(temp_net_income[0])
                b_index += 1
            # 레벨
            if '레벨 업' in daily_data[i][j]:
                level_up_index.append(i + 2)  # 레벨 업 한 날짜 저장
                sheet.cell(row=i + 2, column=6).value = len(level_up_index)  # 레벨 업 한 날짜에 해당 레벨 채우기
                # 그 전 공백까지 직전 레벨로 채우기
                for days in range(level_up_index[len(level_up_index) - 2], level_up_index[len(level_up_index) - 1] - 1):
                    sheet.cell(row=days + 1, column=6).value = len(level_up_index) - 1
                # 테스트용 프린트
                # print(level_up_index[len(level_up_index)-2], ' ', level_up_index[len(level_up_index)-1]-1)
            # 광고, 스킬 사용 레시피등록
            if '광고' in daily_data[i][j] or '랜덤' in daily_data[i][j]:
                if '골드' in daily_data[i][j] and '신문' not in daily_data[i][j]:
                    gold_count += 1
                elif '다이아' in daily_data[i][j] and '신문' not in daily_data[i][j]:
                    dia_count += 1
                elif '피버 타임' in daily_data[i][j] and '끝' not in daily_data[i][j]:
                    fever_count += 1
                elif '러쉬 타임' in daily_data[i][j] and '끝' not in daily_data[i][j]:
                    rush_count += 1
                elif '더블 타임' in daily_data[i][j] and '끝' not in daily_data[i][j]:
                    double_count += 1
                elif '레시피' in daily_data[i][j]:
                    recipe_count += 1
                else:
                    if '물약' in daily_data[i][j]:
                        if not is_random_potion_counted:
                            newspaper_count += 1
                            is_random_potion_counted = True
                        else:
                            continue
                    else:
                        newspaper_count += 1
                ad_count = fever_count + rush_count + double_count + newspaper_count + gold_count + dia_count
            # 물약 사용
            if '포션' in daily_data[i][j]:
                if 'popularity' in daily_data[i][j] and '랜덤' not in daily_data[i][j]:
                    sheet.cell(row=d_index, column=10).value = draw_data
                if 'saving' in daily_data[i][j] and '랜덤' not in daily_data[i][j]:
                    sheet.cell(row=d_index, column=11).value = draw_data
                if 'gold' in daily_data[i][j] and '랜덤' not in daily_data[i][j]:
                    sheet.cell(row=d_index, column=12).value = draw_data
                if 'productivity' in daily_data[i][j] and '랜덤' not in daily_data[i][j]:
                    sheet.cell(row=d_index, column=13).value = draw_data
                if 'speed' in daily_data[i][j] and '랜덤' not in daily_data[i][j]:
                    sheet.cell(row=d_index, column=14).value = draw_data
                if 'attraction' in daily_data[i][j] and '랜덤' not in daily_data[i][j]:
                    sheet.cell(row=d_index, column=15).value = draw_data
                if '랜덤' in daily_data[i][j]:
                    sheet.cell(row=d_index, column=16).value = draw_data
        # ad_count 는 일간으로 출력
        # print(ad_count, ' ', fever_count, ' ', rush_count, ' ', double_count)
        # print(gold_count, ' ', dia_count, ' ', newspaper_count, ' ', recipe_count)
        # ad_count 관련 정보를 outline sheet 에 적용
        sheet.cell(row=c_index, column=7).value = fever_count
        sheet.cell(row=c_index, column=8).value = rush_count
        sheet.cell(row=c_index, column=9).value = double_count
        sheet.cell(row=c_index, column=17).value = ad_count
        c_index += 1
        d_index += 1  # 포션 사용은 안 하는 경우도 있기 때문에 일자 끝난 뒤 인덱스 추가


# 시트 1(store) 데이터를 채움
def fill_store_sheet(sheet, lines):
    # 매장 영업 중 로그에서 손님 입장, 주문, 피드백, 결제 찾아서 채우기
    # 피드백은 괄호 안의 숫자 (n) 로 찾기
    customer_data = []
    cut_line_by_entrance_and_exit(lines, customer_data)

    index = 2  # 일자 상관 없이 손님 수에 따라 +1
    # print(customer_data)  # 손님 데이터 잘 잘렸는지 테스트

    for i in range(len(customer_data)):
        # 각 손님에 대해 데이터가 필요한 로그의 인덱스 받아와서 딕셔너리에 저장
        customer_log_index_dict = {}
        customer_key = customer_data[i][0].split(':')[3].split()[0]
        for j in range(len(customer_data[i])):
            if '카페 서빙' in customer_data[i][j] and '피버타임' not in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_cafe_serve'] = j
            elif '브런치 서빙' in customer_data[i][j] and '피버타임' not in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_brunch_serve'] = j
            elif '피버타임 카페' in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_fever_cafe_serve'] = j
            elif '피버타임 브런치' in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_fever_brunch_serve'] = j
            elif '조합' in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_purchase'] = j
            elif '미니스토리 시작' in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_mini_story_start'] = j
            elif '미니스토리 정답' in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_mini_story_success'] = j
            elif '미니스토리 오답' in customer_data[i][j]:
                customer_log_index_dict[customer_key + '_mini_story_fail'] = j
            else:
                continue
        # print(customer_log_index_dict)

        # 이하 받아온 인덱스 값을 기반으로 셀 내용 채우기
        # 관련 변수들
        cafe_serve_menu = ''  # 카페 서빙 메뉴
        brunch_serve_menu = ''  # 브런치 서빙 메뉴
        cafe_feedback = 0  # 카페 피드백 인덱스
        brunch_feedback = 0  # 브런치 피드백 인덱스

        # 손님 동물종
        sheet.cell(row=index, column=2).value = customer_key
        if customer_log_index_dict.get(customer_key + '_cafe_serve'):
            cafe_serve_menu = customer_data[i][customer_log_index_dict[customer_key + '_cafe_serve']].split('(')[1].split(')')[0]
            cafe_feedback = int(
                customer_data[i][customer_log_index_dict[customer_key + '_cafe_serve'] + 1].split('(')[1].split(')')[0])
        elif customer_log_index_dict.get(customer_key + '_fever_cafe_serve'):
            cafe_serve_menu = 'F'
        # 브런치 서빙
        if customer_log_index_dict.get(customer_key + '_brunch_serve'):
            brunch_serve_menu = customer_data[i][customer_log_index_dict[customer_key + '_brunch_serve']].split('(')[1].split(')')[0]
            brunch_feedback = int(
                customer_data[i][customer_log_index_dict[customer_key + '_brunch_serve'] + 1].split('(')[1].split(')')[
                    0])
        elif customer_log_index_dict.get(customer_key + '_fever_brunch_serve'):
            brunch_serve_menu = 'F'
        # 미니스토리와 조합
        if customer_log_index_dict.get(customer_key + '_mini_story_start'):
            if customer_log_index_dict.get(customer_key + '_mini_story_success'):
                sheet.cell(row=index, column=8).value = '정답'
                # print(customer_key, ' ', '미니스토리 정답', ' ', index)
            elif customer_log_index_dict.get(customer_key + '_mini_story_fail'):
                sheet.cell(row=index, column=8).value = '오답'
                # print(customer_key, ' ', '미니스토리 오답', ' ', index)
        else:
            purchase_feedback = int(customer_data[i][customer_log_index_dict[customer_key + '_purchase']].split('(')[1].split(')')[0])
            sheet.cell(row=index, column=3).value = cafe_serve_menu
            sheet.cell(row=index, column=4).value = cafe_feedback
            sheet.cell(row=index, column=5).value = brunch_serve_menu
            sheet.cell(row=index, column=6).value = brunch_feedback
            sheet.cell(row=index, column=7).value = purchase_feedback
            # print(customer_key + ' ' + cafe_serve_menu + ' ' + str(cafe_feedback) + ' ' + brunch_serve_menu + ' ' + str(brunch_feedback) + ' ' + str(purchase_feedback))

        index += 1  # 한 손님의 로그 처리 완료


customer_info = {}  # key 값과 방문일자 저장


# 시트 2(customer) 데이터를 채움
def fill_customer_sheet(sheet, lines):
    # store sheet 로그 분석에서 자른 내용 중 단골 손님과 관련된 부분의 cafe, brunch 피드백만 가지고 내용 채우기
    # 피드백은 괄호 안의 숫자 (n) 로 찾기
    customer_data = []
    cut_line_by_entrance_and_exit(lines, customer_data)
    # print(customer_data[0])

    # print(customer_info)

    # 단골 손님 key 딕셔너리
    special_customer_dic = {'hedgehog': 0, 'cat': 0, 'racoon': 0, 'wolf': 0, 'squirrel': 0, 'mole': 0, 'owl': 0,
                            'ferret': 0}
    for i in range(len(customer_data)):
        customer_log_index_dict = {}
        customer_key = customer_data[i][0].split(':')[3].split()[0]
        if customer_key in special_customer_dic.keys():
            special_customer_dic[customer_key] += 1  # 중복 방지 인덱스 추가
            new_customer_key = customer_key + '_' + str(special_customer_dic[customer_key])
            # print(new_customer_key)
            for j in range(len(customer_data[i])):
                if '카페 서빙' in customer_data[i][j] and '피버타임' not in customer_data[i][j]:
                    customer_log_index_dict[customer_key + '_cafe_serve'] = j
                elif '브런치 서빙' in customer_data[i][j] and '피버타임' not in customer_data[i][j]:
                    customer_log_index_dict[customer_key + '_brunch_serve'] = j
                elif '피버타임 카페' in customer_data[i][j]:
                    customer_log_index_dict[customer_key + '_fever_cafe_serve'] = j
                elif '피버타임 브런치' in customer_data[i][j]:
                    customer_log_index_dict[customer_key + '_fever_brunch_serve'] = j
                else:
                    continue
            # new_customer_key 로 인덱스(Day) 계산
            index = 4 + customer_info[new_customer_key]
            # print(index)
        else:
            continue
        # print(customer_log_index_dict)

        # 미니스토리는 아무 표기 X
        if customer_log_index_dict == {}:
            continue
        else:
            cafe_feedback = 0
            brunch_feedback = 0
            # 카페 피드백
            if customer_log_index_dict.get(customer_key + '_cafe_serve'):
                cafe_feedback = int(
                    customer_data[i][customer_log_index_dict[customer_key + '_cafe_serve'] + 1].split('(')[1].split(')')[0])
            # 브런치 피드백
            if customer_log_index_dict.get(customer_key + '_brunch_serve'):
                brunch_feedback = int(
                    customer_data[i][customer_log_index_dict[customer_key + '_brunch_serve'] + 1].split('(')[1].split(')')[0])
            # print(customer_key + ' ' + str(cafe_feedback) + ' ' + str(brunch_feedback))

            start_feedback_string = str(sheet.cell(row=index, column=list(special_customer_dic).index(customer_key) + 2).value)
            if start_feedback_string == 'None':
                start_feedback_string = ''

            add_feedback_string = ''
            # print(str(cafe_feedback) + '_' + str(brunch_feedback))
            # 성공 시 S, 보통 시 N, 실패 시 F 를 연달아 적기 (연속 방문시에도)
            if 1 <= cafe_feedback <= 4:
                # Success
                add_feedback_string += 'S'
            elif 5 <= cafe_feedback <= 8:
                # Normal
                add_feedback_string += 'N'
            elif 9 <= cafe_feedback:
                # Fail
                add_feedback_string += 'F'

            # 피버타임 카페 피드백
            if customer_log_index_dict.get(customer_key + '_fever_cafe_serve'):
                # Normal
                add_feedback_string += 'N'

            if 1 <= brunch_feedback <= 4:
                # Success
                add_feedback_string += 'S'
            elif 5 <= brunch_feedback <= 8:
                # Normal
                add_feedback_string += 'N'
            elif 9 <= brunch_feedback:
                # Fail
                add_feedback_string += 'F'

            # 피버타임 브런치 피드백
            if customer_log_index_dict.get(customer_key + '_fever_brunch_serve'):
                # Normal
                add_feedback_string += 'N'

            # print(add_feedback_string)
            if start_feedback_string == add_feedback_string:
                # print('스킵')
                continue
            else:
                sheet.cell(row=index, column=list(special_customer_dic).index(customer_key) + 2).value = start_feedback_string + add_feedback_string


# 시트 3(recipe) 데이터를 채움
def fill_recipe_sheet(sheet, lines):
    daily_data = []
    cut_line_by_day(daily_data, lines)

    # '레시피 등록 신청' 체크하여 레시피 key 받아오기
    # '레시피 등록 성공' or '레시피 등록 실패' 로 성공 여부 받아오기
    # 성공 혹은 실패 (초회 발송) 전날 광고 시청했는지 체크하기
    # customer_sheet 에서 purchase(7) column 조사하여 관련 피드백 있었는지 확인하기

    draw_data = 'ㅇ'  # 빈 칸 채우는 동그라미 모양
    base_index = 2  # 기본이 되는 인덱스
    index = 0  # 등록 신청한 레시피 수
    recipe_key_dict = {}  # 등록 신청한 레시피 딕셔너리
    for i in range(len(daily_data)):
        for j in range(len(daily_data[i])):
            if '레시피 등록 신청' in daily_data[i][j]:
                recipe_key = daily_data[i][j].split(':')[3].split()[0]
                # print(recipe_key)
                recipe_key_dict[recipe_key] = [[i, j], 0]
            elif '레시피 등록 성공' in daily_data[i][j]:
                recipe_key = daily_data[i][j].split(':')[3].split()[0]
                temp_dict_data = recipe_key_dict[recipe_key]
                # 해당 키 관련 첫 알람
                if temp_dict_data[1] < 1:
                    temp_dict_data[1] += 1
                    temp_dict_data.append([i, j])
                    temp_dict_data.append(index)
                    sheet.cell(row=base_index+index, column=5).value = draw_data
                    index += 1  # 레시피 하나가 완성된 것으로 친다.
                    # print(recipe_key, " ", temp_dict_data)
                    recipe_key_dict[recipe_key] = temp_dict_data  # ad 확인을 위한 딕셔너리 데이터 갱신
                else:
                    continue
            elif '레시피 등록 실패' in daily_data[i][j]:
                recipe_key = daily_data[i][j].split(':')[3].split()[0]
                temp_dict_data = recipe_key_dict[recipe_key]
                # 해당 키 관련 첫 알람
                if temp_dict_data[1] < 1:
                    temp_dict_data[1] += 1
                    temp_dict_data.append([i, j])
                    temp_dict_data.append(index)
                    sheet.cell(row=base_index+index, column=6).value = draw_data
                    index += 1  # 레시피 하나가 완성된 것으로 친다.
                    # print(recipe_key, " ", temp_dict_data)
                    recipe_key_dict[recipe_key] = temp_dict_data  # ad 확인을 위한 딕셔너리 데이터 갱신
                else:
                    continue
            else:
                continue

    # 결과 안 나온 레시피 버리기
    for key in range(len(recipe_key_dict)):
        temp_key = list(recipe_key_dict.keys())[key]
        if len(recipe_key_dict[temp_key]) < 3:
            del(recipe_key_dict[temp_key])
            key += 1
    # print(recipe_key_dict)

    # 해당 레시피에 대해 광고를 봤는지의 여부를 계산
    for r in range(len(recipe_key_dict)):
        temp_key = list(recipe_key_dict.keys())[r]
        # 레시피 등록 신청일과 레시피 결과 나온 날에 해당하는 추출된 좌표들
        recipe_apply_i = recipe_key_dict[temp_key][0][0]
        recipe_apply_j = recipe_key_dict[temp_key][0][1]
        recipe_result_i = recipe_key_dict[temp_key][2][0]
        recipe_result_j = recipe_key_dict[temp_key][2][1]
        is_recipe_on = False  # 레시피 등록 기간중 여부를 체크
        temp_section = []  # 레시피별 임시 섹션
        for x in range(len(daily_data)):
            for y in range(len(daily_data[x])):
                # 시작 인덱스
                if x == recipe_apply_i and y == recipe_apply_j:
                    is_recipe_on = True
                if is_recipe_on:
                    if '레시피' in daily_data[x][y]:
                        temp_section.append((daily_data[x][y]))
                        # 종료 인덱스
                        if x == recipe_result_i and y == recipe_result_j:
                            is_recipe_on = False
                else:
                    continue

        # 레시피 키 셀 입력 (결과 나오지 않은 것 제거 이후 입력)
        sheet.cell(row=base_index+recipe_key_dict[temp_key][3], column=2).value = temp_key

        # print(temp_section)
        # 광고 시청 했는지 여부 확인
        for line in temp_section:
            if '광고' in line:
                sheet.cell(base_index+recipe_key_dict[temp_key][3], column=3).value = draw_data
                # print(line)
                continue
            else:
                continue

    # 해당 레시피에 대해 플레이 도중 조합을 봤는지에 대한 계산
    for r in range(len(recipe_key_dict)):
        temp_key = list(recipe_key_dict.keys())[r]
        # 레시피 등록 신청일에 해당하는 추출된 좌표들
        recipe_apply_i = recipe_key_dict[temp_key][0][0]
        recipe_apply_j = recipe_key_dict[temp_key][0][1]
        is_recipe_off = True  # 레시피 등록 신청 직전까지의 로그를 체크
        served_menu_list = []  # 플레이 도중 서빙된 메뉴 리스트
        for a in range(len(daily_data)):
            for b in range(len(daily_data[a])):
                if is_recipe_off:
                    if '카페 서빙' in daily_data[a][b] and '피버타임' not in daily_data[a][b]:
                        cafe_key = daily_data[a][b].split('(')[1].split(')')[0].split()[0]
                        served_menu_list.append(cafe_key)
                    elif '브런치 서빙' in daily_data[a][b] and '피버타임' not in daily_data[a][b]:
                        brunch_key = daily_data[a][b].split('(')[1].split(')')[0].split()[0]
                        served_menu_list.append(brunch_key)
                    # 종료 인덱스
                    if a == recipe_apply_i and b == recipe_apply_j:
                        is_recipe_off = False
                else:
                    continue
        # print(served_menu_list)
        if temp_key in served_menu_list:
            sheet.cell(base_index+recipe_key_dict[temp_key][3], column=4).value = draw_data


# 시트 4 (quest) 데이터를 채움
def fill_quest_sheet(sheet, lines):
    # '일일 퀘스트 달성' 로그 찾아서 해당 칸 채우기
    daily_data = []
    cut_line_by_day(daily_data, lines)

    draw_data = 'ㅇ'
    index = 2  # 일자에 따라 더해짐
    for i in range(len(daily_data)):
        for j in range(len(daily_data[i])):
            if '일일 퀘스트' in daily_data[i][j]:
                quest_key = daily_data[i][j].split(':')[3].split()[0]
                # print(quest_key)
                if quest_key == 'skill_1':
                    sheet.cell(row=index, column=2).value = draw_data
                elif quest_key == 'skill_3':
                    sheet.cell(row=index, column=3).value = draw_data
                elif quest_key == 'ad_1':
                    sheet.cell(row=index, column=4).value = draw_data
                elif quest_key == 'ad_3':
                    sheet.cell(row=index, column=5).value = draw_data
                elif quest_key == 'latte_1':
                    sheet.cell(row=index, column=6).value = draw_data
                elif quest_key == 'latte_2':
                    sheet.cell(row=index, column=7).value = draw_data
                elif quest_key == 'ade_1':
                    sheet.cell(row=index, column=8).value = draw_data
                elif quest_key == 'ade_2':
                    sheet.cell(row=index, column=9).value = draw_data
                elif quest_key == 'juice_1':
                    sheet.cell(row=index, column=10).value = draw_data
                elif quest_key == 'juice_2':
                    sheet.cell(row=index, column=11).value = draw_data
                elif quest_key == 'sandwich_1':
                    sheet.cell(row=index, column=12).value = draw_data
                elif quest_key == 'sandwich_2':
                    sheet.cell(row=index, column=13).value = draw_data
                elif quest_key == 'pancake_1':
                    sheet.cell(row=index, column=14).value = draw_data
                elif quest_key == 'pancake_2':
                    sheet.cell(row=index, column=15).value = draw_data
                elif quest_key == 'heart_1':
                    sheet.cell(row=index, column=16).value = draw_data
                elif quest_key == 'heart_3':
                    sheet.cell(row=index, column=17).value = draw_data
                elif quest_key == 'wash_1':
                    sheet.cell(row=index, column=18).value = draw_data
                elif quest_key == 'wash_3':
                    sheet.cell(row=index, column=19).value = draw_data
                elif quest_key == 'ingredient_1':
                    sheet.cell(row=index, column=20).value = draw_data
                elif quest_key == 'ingredient_3':
                    sheet.cell(row=index, column=21).value = draw_data
                else:
                    print('퀘스트 키 에러', daily_data[i][j])
        index += 1


# (시트 1 - store 분석용) daily_data 를 손님 별로 자르기
def cut_line_by_entrance_and_exit(lines, section):
    # 중간에 다른 손님의 내용이 껴 있다면 섹션에서 내용을 뺀다.
    daily_data = []  # 로그를 일자별로 나눈 것
    cut_line_by_day(daily_data, lines)

    customer_dic = {'rabbit': 0, 'whiterabbit': 0, 'fox': 0, 'redfox': 0, 'whitefox': 0, 'bear': 0, 'whitebear': 0,
                    'blackbear': 0, 'owl': 0, 'squirrel': 0, 'wolf': 0, 'cat': 0, 'mole': 0, 'ferret': 0, 'hedgehog': 0,
                    'racoon': 0}  # 두루미 방문은 제외됨
    customer_entrance_exit_dictionary = {}
    for i in range(len(daily_data)):
        visited_customer_list = []  # 일간 방문한 손님 리스트
        for j in range(len(daily_data[i])):
            # 손님 입장, 손님 퇴장 줄의 인덱스 번호를 가져와서 리스트에 저장
            if '입장' in daily_data[i][j]:
                # 일자 데이터 중 손님 key 값들 추출
                customer_key = daily_data[i][j].split(':')[3].split()[0]
                # print(customer_key)
                customer_dic[customer_key] += 1  # 중복 방지 인덱스 추가
                new_customer_key = customer_key + '_' + str(customer_dic[customer_key])
                # 손님 종이 이미 방문했을 경우 (전체 플레이 데이터를 통틀어서)
                if new_customer_key in customer_entrance_exit_dictionary:
                    # 딕셔너리에 '입장' 인덱스 값 넣기 (인덱스 1부터 시작)
                    customer_entrance_exit_dictionary[new_customer_key] = [[i, j], False]
                    visited_customer_list.append(new_customer_key)
                    return_customer_key_and_visited_day(new_customer_key, i)
                # 손님 종이 처음 방문했을 경우 (전체 플레이 데이터를 통틀어서)
                else:
                    visited_customer_list.append(new_customer_key)
                    # 딕셔너리에 '입장' 인덱스 값 넣기
                    customer_entrance_exit_dictionary[new_customer_key] = [[i, j], False]
                    return_customer_key_and_visited_day(new_customer_key, i)
            elif '퇴장' in daily_data[i][j]:
                customer_key = daily_data[i][j].split(':')[3].split()[0]
                new_customer_key = customer_key + '_' + str(customer_dic[customer_key])
                # '퇴장' 로그가 먼저 나올 경우 (좌석이 없어 바로 퇴장할 경우)
                if new_customer_key not in customer_entrance_exit_dictionary:
                    # print(new_customer_key)
                    customer_entrance_exit_dictionary[new_customer_key] = 'destroy'
                # 일반적인 경우
                else:
                    # '퇴장' 로그가 여러 번 출력되는 경우가 있어 한 번만 인식
                    if not customer_entrance_exit_dictionary[new_customer_key][1]:
                        temp_values = customer_entrance_exit_dictionary[new_customer_key]
                        temp_values.append([i, j])
                        customer_entrance_exit_dictionary[new_customer_key] = temp_values
                        customer_entrance_exit_dictionary[new_customer_key][1] = True
                        # print(new_customer_key, ' ', customer_entrance_exit_dictionary[new_customer_key][1])
                    else:
                        continue
            else:
                continue
        # print(visited_customer_list)  # 일간 방문 손님 리스트를 출력

    # print(len(customer_entrance_exit_dictionary))  # 최종 딕셔너리에서 나온 값들의 개수 출력
    # 잘못된 값들 삭제
    dict_keys = list(customer_entrance_exit_dictionary.keys())
    for value in range(len(customer_entrance_exit_dictionary)):
        if len(customer_entrance_exit_dictionary[dict_keys[value]]) != 3:
            # print('wrong')
            del[customer_entrance_exit_dictionary[dict_keys[value]]]
            value += 1
        else:
            # print(customer_entrance_exit_dictionary[dict_keys[value]])
            continue
    # print(len(customer_entrance_exit_dictionary))  # 잘못 삭제되진 않았는지 체크

    # 딕셔너리의 키(손님) 각각에 대해 해당하는 섹션을 만들어준다.
    # 섹션 내부에 관련 없는 내용이 있을 경우 해당 내용을 섹션에서 제외시킨다.
    for t in range(len(customer_entrance_exit_dictionary)):
        temp_key = list(customer_entrance_exit_dictionary.keys())[t]
        # print(temp_key, ' ', customer_entrance_exit_dictionary[temp_key])
        # 입장과 퇴장에 해당하는 추출된 좌표들
        entrance_i = customer_entrance_exit_dictionary[temp_key][0][0]
        entrance_j = customer_entrance_exit_dictionary[temp_key][0][0]
        exit_i = customer_entrance_exit_dictionary[temp_key][2][0]
        exit_j = customer_entrance_exit_dictionary[temp_key][2][1]
        is_entrance_on = False  # 입장 여부를 체크
        temp_section = []  # 임시 섹션
        for x in range(len(daily_data)):
            for y in range(len(daily_data[x])):
                # 시작 인덱스
                if x == entrance_i and y == entrance_j:
                    is_entrance_on = True
                    temp_section.append(daily_data[x][y])
                if is_entrance_on:
                    customer_key = temp_key.split('_')[0]
                    # 섹션 내부에 관련 있는 내용만 남긴다.
                    if customer_key in daily_data[x][y] and check_the_name_is_same(customer_key, daily_data[x][y]):
                        temp_section.append(daily_data[x][y])
                        # 종료 인덱스
                        if x == exit_i and y == exit_j:
                            is_entrance_on = False
                    else:
                        continue

        # print(list(customer_entrance_exit_dictionary.keys())[t], " ", temp_section)

        # temp_section 에서 같은 동물종이 여러 번 방문할 경우 데이터가 누적되는 문제 해결
        # 섹션 안에 '입장' 이 여럿인지 체크한다.
        check_single_count = 0
        multiple_entrance_index = []
        for d in range(len(temp_section)):
            if '입장' in temp_section[d]:
                check_single_count += 1
                multiple_entrance_index.append(d)

        # '입장' 이 여럿일 경우 혹은 첫 인덱스가 '입장' 이 아닐 경우
        # temp_section 안에서 마지막 '입장' 을 찾아 그 앞을 자른다.
        if check_single_count > 1 or '입장' not in temp_section[0]:
            cut_temp_section = temp_section[multiple_entrance_index[len(multiple_entrance_index) - 1]:]
            if len(cut_temp_section) > 4:
                section.append(cut_temp_section)  # 최소 = len(미니스토리) 가 5
            # 잘못 인식된 것들
            else:
                continue
        else:
            if len(temp_section) > 4:
                section.append(temp_section)
            # 잘못 인식된 것들
            else:
                continue

    # 최종 section 출력
    # for line in section:
        # print(line)
    # print(section)


def return_customer_key_and_visited_day(customer_key, visited_day):
    customer_info[customer_key] = visited_day
    return


def check_the_name_is_same(name, contents):
    if name == 'bear':
        if 'whitebear' in contents:
            return False
        elif 'blackbear' in contents:
            return False
        else:
            return True
    elif name == 'fox':
        if 'whitefox' in contents:
            return False
        elif 'redfox' in contents:
            return False
        else:
            return True
    elif name == 'rabbit':
        if 'whiterabbit' in contents:
            return False
        else:
            return True
    else:
        return True


# (Day 별 분석용) lines 를 Day 별로 자르기
def cut_line_by_day(section, lines):
    # ===== StartGame 이나 Loading ===== 줄의 인덱스 번호를 가져와서 리스트에 저장
    day_cut_line_index = []
    for i in range(len(lines)):
        if 'StartGame' in lines[i] or 'LoadingScene' in lines[i]:
            day_cut_line_index.append(i)
        else:
            continue

    # 인덱스 번호에 따라 섹션으로 자름
    for j in range(len(day_cut_line_index) - 1):
        section.append(lines[day_cut_line_index[j]:day_cut_line_index[j + 1]])

def write_shop_data_xlsx(user_name):
    from openpyxl import load_workbook

    shop_data_file = load_workbook('_files/' + user_name + '/shopdata.xlsx')

    ingredient_sheet = shop_data_file['ingredient']
    equipment_sheet = shop_data_file['equipment']
    potion_sheet = shop_data_file['potion']
    interior_wallfloor_sheet = shop_data_file['interior_wallfloor']
    interior_structure_sheet = shop_data_file['interior_structure']
    interior_store_sheet = shop_data_file['interior_store']
    interior_kitchen_sheet = shop_data_file['interior_kitchen']
    interior_deco_sheet = shop_data_file['interior_deco']
    interior_sheet_list = [interior_wallfloor_sheet, interior_structure_sheet, interior_store_sheet, interior_kitchen_sheet, interior_deco_sheet]

    # 파일 열고 읽기
    with open('_files/' + user_name + '/Log - ' + user_name + '.txt', encoding='utf-16') as f:
        lines = f.readlines()

    # 셀 내용 채우기 함수 실행
    fill_ingredient_sheet(ingredient_sheet, lines)
    fill_equipment_sheet(equipment_sheet, lines)
    fill_potion_sheet(potion_sheet, lines)
    fill_interior_sheets(interior_sheet_list, lines)

    # 파일 저장 및 닫기
    shop_data_file.save(filename='_files/' + user_name + '/shopdata.xlsx')
    shop_data_file.close()

    # 끝났는지 확인
    print(user_name, 'shop_data done')


# 시트 1(ingredient) 데이터를 채움
def fill_ingredient_sheet(sheet, lines):
    daily_data = []
    cut_line_by_day(daily_data, lines)

    draw_data = 'ㅇ'
    index = 2  # 일자에 따라 더해짐
    for i in range(len(daily_data)):
        for j in range(len(daily_data[i])):
            if '재료 오픈' in daily_data[i][j]:
                ingredient_key = daily_data[i][j].split(':')[3].split()[0]
                # print(ingredient_key)
                if ingredient_key == '꿀':
                    sheet.cell(row=index, column=2).value = draw_data
                elif ingredient_key == '치즈':
                    sheet.cell(row=index, column=3).value = draw_data
                elif ingredient_key == '바나나':
                    sheet.cell(row=index, column=4).value = draw_data
                elif ingredient_key == '버터':
                    sheet.cell(row=index, column=5).value = draw_data
                elif ingredient_key == '녹차':
                    sheet.cell(row=index, column=6).value = draw_data
                elif ingredient_key == '레몬':
                    sheet.cell(row=index, column=7).value = draw_data
                elif ingredient_key == '당근':
                    sheet.cell(row=index, column=8).value = draw_data
                elif ingredient_key == '토마토':
                    sheet.cell(row=index, column=9).value = draw_data
                elif ingredient_key == '땅콩':
                    sheet.cell(row=index, column=10).value = draw_data
                elif ingredient_key == '포도':
                    sheet.cell(row=index, column=11).value = draw_data
                elif ingredient_key == '자몽':
                    sheet.cell(row=index, column=12).value = draw_data
                elif ingredient_key == '복숭아':
                    sheet.cell(row=index, column=13).value = draw_data
                elif ingredient_key == '옥수수':
                    sheet.cell(row=index, column=14).value = draw_data
                elif ingredient_key == '체리':
                    sheet.cell(row=index, column=15).value = draw_data
                elif ingredient_key == '오이':
                    sheet.cell(row=index, column=16).value = draw_data
                elif ingredient_key == '호박':
                    sheet.cell(row=index, column=17).value = draw_data
                elif ingredient_key == '아몬드':
                    sheet.cell(row=index, column=18).value = draw_data
                else:
                    print('키값 오류', ingredient_key)
            else:
                continue
        index += 1


# 시트 2(equipment) 데이터를 채움
def fill_equipment_sheet(sheet, lines):
    daily_data = []
    cut_line_by_day(daily_data, lines)

    draw_data = 'ㅇ'
    index = 2  # 일자에 따라 더해짐
    for i in range(len(daily_data)):
        for j in range(len(daily_data[i])):
            if '설비 업그레이드 구매' in daily_data[i][j]:
                equipment_key = daily_data[i][j].split(':')[3].split('_')[0].split()[0]
                if equipment_key == 'table':
                    sheet.cell(row=index, column=2).value = draw_data
                elif equipment_key == 'brunchSet':
                    sheet.cell(row=index, column=3).value = draw_data
                elif equipment_key == 'pan':
                    sheet.cell(row=index, column=4).value = draw_data
                elif equipment_key == 'oven':
                    sheet.cell(row=index, column=5).value = draw_data
                elif equipment_key == 'board':
                    sheet.cell(row=index, column=6).value = draw_data
                elif equipment_key == 'servingList':
                    sheet.cell(row=index, column=7).value = draw_data
                elif equipment_key == 'refrigerator':
                    sheet.cell(row=index, column=8).value = draw_data
                elif equipment_key == 'smartMachine':
                    sheet.cell(row=index, column=9).value = draw_data
                elif equipment_key == 'smartBoard':
                    sheet.cell(row=index, column=10).value = draw_data
                elif equipment_key == 'smartSink':
                    sheet.cell(row=index, column=11).value = draw_data
                elif equipment_key == 'smartDrink':
                    sheet.cell(row=index, column=12).value = draw_data
                else:
                    print('키 값 에러', equipment_key)
            else:
                continue
        index += 1


# 시트 3(potion) 데이터를 채움
def fill_potion_sheet(sheet, lines):
    daily_data = []
    cut_line_by_day(daily_data, lines)

    index = 2  # 일자에 따라 더해짐
    for i in range(len(daily_data)):
        # 일자별로 물약, 비약, 순백의 비약 구매 개수 초기화 (첫 구매 시 1을 적음)
        # 일자별 구매 개수를 셀에 적는다.
        popularity_count = 1
        saving_count = 1
        gold_count = 1
        productivity_count = 1
        speed_count = 1
        attraction_count = 1
        random_count = 1
        s_count_count = 1
        s_gold_count = 1
        s_popularity_count = 1
        s_speed_count = 1
        s_productivity_count = 1
        s_hurry_count = 1
        for j in range(len(daily_data[i])):
            if '물약 구매' in daily_data[i][j]:
                potion_key = daily_data[i][j].split('$$$')[1].split('물약')[0].split()[0]
                # print('물약', potion_key)
                if potion_key == 'popularity':
                    sheet.cell(row=index, column=2).value = popularity_count
                    popularity_count += 1
                elif potion_key == 'saving':
                    sheet.cell(row=index, column=3).value = saving_count
                    saving_count += 1
                elif potion_key == 'gold':
                    sheet.cell(row=index, column=4).value = gold_count
                    gold_count += 1
                elif potion_key == 'productivity':
                    sheet.cell(row=index, column=5).value = productivity_count
                    productivity_count += 1
                elif potion_key == 'speed':
                    sheet.cell(row=index, column=6).value = speed_count
                    speed_count += 1
                elif potion_key == 'attraction':
                    sheet.cell(row=index, column=7).value = attraction_count
                    attraction_count += 1
                elif potion_key == 'random':
                    sheet.cell(row=index, column=8).value = random_count
                    random_count += 1
                else:
                    print('물약 키 값 에러', potion_key)
            elif '순백의 비약 구매' in daily_data[i][j]:
                # potion_key = 'sCount'
                # print('순백', potion_key)
                sheet.cell(row=index, column=9).value = s_count_count
                s_count_count += 1
            elif '비약 구매' in daily_data[i][j]:
                potion_key = daily_data[i][j].split('$$$')[1].split('비약')[0].split()[0]
                # print('비약', potion_key)
                if potion_key == 'sGold':
                    sheet.cell(row=index, column=10).value = s_gold_count
                    s_gold_count += 1
                elif potion_key == 'sPopularity':
                    sheet.cell(row=index, column=11).value = s_popularity_count
                    s_popularity_count += 1
                elif potion_key == 'sSpeed':
                    sheet.cell(row=index, column=12).value = s_speed_count
                    s_speed_count += 1
                elif potion_key == 'sProductivity':
                    sheet.cell(row=index, column=13).value = s_productivity_count
                    s_productivity_count += 1
                elif potion_key == 'sHurry':
                    sheet.cell(row=index, column=14).value = s_hurry_count
                    s_hurry_count += 1
                else:
                    print('비약 키 값 에러', potion_key)
            else:
                continue
        index += 1


# 시트 4~8(interior_) 데이터를 채움
def fill_interior_sheets(sheets, lines):
    daily_data = []
    cut_line_by_day(daily_data, lines)

    draw_data = 'ㅇ'
    index = 3  # 일자에 따라 더해짐
    for i in range(len(daily_data)):
        for j in range(len(daily_data[i])):
            if '인테리어 구매' in daily_data[i][j]:
                interior_part = daily_data[i][j].split(':')[3].split('_')[0].split()[0]
                interior_key = daily_data[i][j].split(':')[3].split('_')[1].split()[0]
                print(interior_part, interior_key)
                # wallfloor(sheet[0]) 시트
                if interior_part == 'wall':
                    if interior_key == 'spring1':
                        sheets[0].cell(row=index, column=2).value = draw_data
                    elif interior_key == 'spring2':
                        sheets[0].cell(row=index, column=3).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[0].cell(row=index, column=4).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[0].cell(row=index, column=5).value = draw_data
                    elif interior_key == 'fall':
                        sheets[0].cell(row=index, column=6).value = draw_data
                    elif interior_key == 'winter1':
                        sheets[0].cell(row=index, column=7).value = draw_data
                    elif interior_key == 'winter2':
                        sheets[0].cell(row=index, column=8).value = draw_data
                    elif interior_key == 'sen1':
                        sheets[0].cell(row=index, column=9).value = draw_data
                    elif interior_key == 'sen2':
                        sheets[0].cell(row=index, column=10).value = draw_data
                    else:
                        print('인테리어 wall 키 값 에러')
                elif interior_part == 'floor':
                    if interior_key == 'spring1':
                        sheets[0].cell(row=index, column=11).value = draw_data
                    elif interior_key == 'spring2':
                        sheets[0].cell(row=index, column=12).value = draw_data
                    elif interior_key == 'spring3':
                        sheets[0].cell(row=index, column=13).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[0].cell(row=index, column=14).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[0].cell(row=index, column=15).value = draw_data
                    elif interior_key == 'fall':
                        sheets[0].cell(row=index, column=16).value = draw_data
                    elif interior_key == 'winter1':
                        sheets[0].cell(row=index, column=17).value = draw_data
                    elif interior_key == 'winter2':
                        sheets[0].cell(row=index, column=18).value = draw_data
                    else:
                        print('인테리어 floor 키 값 에러')
                elif interior_part == 'kitchen':
                    if interior_key == 'spring':
                        sheets[0].cell(row=index, column=19).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[0].cell(row=index, column=20).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[0].cell(row=index, column=21).value = draw_data
                    elif interior_key == 'fall':
                        sheets[0].cell(row=index, column=22).value = draw_data
                    elif interior_key == 'winter1':
                        sheets[0].cell(row=index, column=23).value = draw_data
                    elif interior_key == 'winter2':
                        sheets[0].cell(row=index, column=24).value = draw_data
                    elif interior_key == 'sen':
                        sheets[0].cell(row=index, column=25).value = draw_data
                    else:
                        print('인테리어 kitchen 키 값 에러')
                # structure(sheet[1]) 시트
                elif interior_part == 'pillow':
                    if interior_key == 'spring':
                        sheets[1].cell(row=index, column=2).value = draw_data
                    elif interior_key == 'summer':
                        sheets[1].cell(row=index, column=3).value = draw_data
                    elif interior_key == 'fall':
                        sheets[1].cell(row=index, column=4).value = draw_data
                    elif interior_key == 'winter':
                        sheets[1].cell(row=index, column=5).value = draw_data
                    else:
                        print('인테리어 pillow 키 값 에러')
                elif interior_part == 'door':
                    if interior_key == 'spring':
                        sheets[1].cell(row=index, column=6).value = draw_data
                    elif interior_key == 'summer':
                        sheets[1].cell(row=index, column=7).value = draw_data
                    elif interior_key == 'fall':
                        sheets[1].cell(row=index, column=8).value = draw_data
                    elif interior_key == 'winter':
                        sheets[1].cell(row=index, column=9).value = draw_data
                    elif interior_key == 'sen':
                        sheets[1].cell(row=index, column=10).value = draw_data
                    else:
                        print('인테리어 door 키 값 에러')
                elif interior_part == 'window':
                    if interior_key == 'spring':
                        sheets[1].cell(row=index, column=11).value = draw_data
                    elif interior_key == 'summer':
                        sheets[1].cell(row=index, column=12).value = draw_data
                    elif interior_key == 'fall':
                        sheets[1].cell(row=index, column=13).value = draw_data
                    elif interior_key == 'winter':
                        sheets[1].cell(row=index, column=14).value = draw_data
                    else:
                        print('인테리어 window 키 값 에러')
                # store(sheet[2]) 시트
                elif interior_part == 'table':
                    if interior_key == 'spring':
                        sheets[2].cell(row=index, column=2).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[2].cell(row=index, column=3).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[2].cell(row=index, column=4).value = draw_data
                    elif interior_key == 'fall':
                        sheets[2].cell(row=index, column=5).value = draw_data
                    elif interior_key == 'winter':
                        sheets[2].cell(row=index, column=6).value = draw_data
                    else:
                        print('인테리어 table 키 값 에러')
                elif interior_part == 'tabletop':
                    if interior_key == 'spring':
                        sheets[2].cell(row=index, column=7).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[2].cell(row=index, column=8).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[2].cell(row=index, column=9).value = draw_data
                    elif interior_key == 'fall':
                        sheets[2].cell(row=index, column=10).value = draw_data
                    elif interior_key == 'winter':
                        sheets[2].cell(row=index, column=11).value = draw_data
                    else:
                        print('인테리어 tabletop 키 값 에러')
                elif interior_part == 'chair':
                    if interior_key == 'spring':
                        sheets[2].cell(row=index, column=12).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[2].cell(row=index, column=13).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[2].cell(row=index, column=14).value = draw_data
                    elif interior_key == 'fall':
                        sheets[2].cell(row=index, column=15).value = draw_data
                    elif interior_key == 'winter':
                        sheets[2].cell(row=index, column=16).value = draw_data
                    else:
                        print('인테리어 chair 키 값 에러')
                elif interior_part == 'bartable':
                    if interior_key == 'spring':
                        sheets[2].cell(row=index, column=17).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[2].cell(row=index, column=18).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[2].cell(row=index, column=19).value = draw_data
                    elif interior_key == 'fall':
                        sheets[2].cell(row=index, column=20).value = draw_data
                    elif interior_key == 'winter':
                        sheets[2].cell(row=index, column=21).value = draw_data
                    else:
                        print('인테리어 bartable 키 값 에러')
                # kitchen(sheet[3]) 시트
                elif interior_part == 'boardtable':
                    if interior_key == 'spring':
                        sheets[3].cell(row=index, column=2).value = draw_data
                    elif interior_key == 'summer':
                        sheets[3].cell(row=index, column=3).value = draw_data
                    elif interior_key == 'fall':
                        sheets[3].cell(row=index, column=4).value = draw_data
                    elif interior_key == 'winter':
                        sheets[3].cell(row=index, column=5).value = draw_data
                    else:
                        print('인테리어 boardtable 키 값 에러')
                elif interior_part == 'distable':
                    if interior_key == 'spring':
                        sheets[3].cell(row=index, column=6).value = draw_data
                    elif interior_key == 'summer':
                        sheets[3].cell(row=index, column=7).value = draw_data
                    elif interior_key == 'fall':
                        sheets[3].cell(row=index, column=8).value = draw_data
                    elif interior_key == 'winter':
                        sheets[3].cell(row=index, column=9).value = draw_data
                    else:
                        print('인테리어 distable 키 값 에러')
                elif interior_part == 'servingtable':
                    if interior_key == 'spring':
                        sheets[3].cell(row=index, column=10).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[3].cell(row=index, column=11).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[3].cell(row=index, column=12).value = draw_data
                    elif interior_key == 'fall':
                        sheets[3].cell(row=index, column=13).value = draw_data
                    elif interior_key == 'winter':
                        sheets[3].cell(row=index, column=14).value = draw_data
                    else:
                        print('인테리어 servingtable 키 값 에러')
                elif interior_part == 'countertable':
                    if interior_key == 'spring':
                        sheets[3].cell(row=index, column=15).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[3].cell(row=index, column=16).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[3].cell(row=index, column=17).value = draw_data
                    elif interior_key == 'fall':
                        sheets[3].cell(row=index, column=18).value = draw_data
                    elif interior_key == 'winter':
                        sheets[3].cell(row=index, column=19).value = draw_data
                    else:
                        print('인테리어 countertable 키 값 에러')
                # deco(sheet[4]) 시트
                elif interior_part == 'carpet':
                    if interior_key == 'spring':
                        sheets[4].cell(row=index, column=2).value = draw_data
                    elif interior_key == 'summer1':
                        sheets[4].cell(row=index, column=3).value = draw_data
                    elif interior_key == 'summer2':
                        sheets[4].cell(row=index, column=4).value = draw_data
                    elif interior_key == 'fall':
                        sheets[4].cell(row=index, column=5).value = draw_data
                    elif interior_key == 'winter':
                        sheets[4].cell(row=index, column=6).value = draw_data
                    else:
                        print('인테리어 carpet 키 값 에러')
                elif interior_part == 'leftobj':
                    if interior_key == 'spring':
                        sheets[4].cell(row=index, column=7).value = draw_data
                    elif interior_key == 'fall1':
                        sheets[4].cell(row=index, column=8).value = draw_data
                    elif interior_key == 'fall2':
                        sheets[4].cell(row=index, column=9).value = draw_data
                    elif interior_key == 'winter':
                        sheets[4].cell(row=index, column=10).value = draw_data
                    else:
                        print('인테리어 leftobj 키 값 에러')
                elif interior_part == 'wallobj':
                    if interior_key == 'spring':
                        sheets[4].cell(row=index, column=11).value = draw_data
                    elif interior_key == 'summer':
                        sheets[4].cell(row=index, column=12).value = draw_data
                    elif interior_key == 'sen':
                        sheets[4].cell(row=index, column=13).value = draw_data
                    else:
                        print('인테리어 wallobj 키 값 에러')
                elif interior_part == 'trashcan':
                    if interior_key == 'green':
                        sheets[4].cell(row=index, column=14).value = draw_data
                    elif interior_key == 'model':
                        sheets[4].cell(row=index, column=15).value = draw_data
                    else:
                        print('인테리어 trashcan 키 값 에러')
                else:
                    print('인테리어 파트 에러', interior_part, interior_key)
            else:
                continue
        index += 1


# (Day 별 분석용) lines 를 Day 별로 자르기
def cut_line_by_day(section, lines):
    # ===== StartGame 이나 Loading ===== 줄의 인덱스 번호를 가져와서 리스트에 저장
    day_cut_line_index = []
    for i in range(len(lines)):
        if 'StartGame' in lines[i] or 'LoadingScene' in lines[i]:
            day_cut_line_index.append(i)
        else:
            continue

    # 인덱스 번호에 따라 섹션으로 자름
    for j in range(len(day_cut_line_index) - 1):
        section.append(lines[day_cut_line_index[j]:day_cut_line_index[j + 1]])

from openpyxl import load_workbook
import game_data_to_xlsx as gd
import shop_data_to_xlsx as sd
analysis_file = load_workbook('_files/analysis.xlsx')

player_key_list = ['샘플']


# 리스트 안에 있는 각각의 key 값에 대한 초회 검사인지 체크
# analysis 파일의 시트 0(index) 데이터를 채우고 game_data_xlsx 와 shop_data_xlsx 를 실행시킴
def check_is_first_for_files(keys):
    index_sheet = analysis_file['index']
    for key in range(len(keys)):
        # row 검사 : 초회
        if str(index_sheet.cell(row=key+1, column=1).value) == 'None':
            index_sheet.cell(row=key+1, column=1).value = keys[key]  # 이름 입력
            analysis_file.save(filename='_files/analysis.xlsx')  # 파일 저장
            analysis_file.close()  # 파일 닫기
            # game_data 쓰기
            gd.write_game_data_xlsx(keys[key])
            # shop_data 쓰기
            sd.write_shop_data_xlsx(keys[key])
        # row 검사 : 초회 이후 -> 스킵
        else:
            print(keys[key], '스킵')
            continue
    return True


check_is_first_for_files(player_key_list)